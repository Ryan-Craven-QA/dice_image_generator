import numpy as np
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook

# Constants
INCHES_PER_FOOT = 12

# Create the main window
root = tk.Tk()
root.title("Dice Image Generator")

# Set minimum size and make the window resizable
root.minsize(800, 600)
root.resizable(True, True)

# Create the main frame
main_frame = ttk.Frame(root)
main_frame.pack(fill=tk.BOTH, expand=True)

# Variables to store user inputs
physical_width_var = tk.DoubleVar()
physical_height_var = tk.DoubleVar()
dice_type_var = tk.StringVar(value='Monochrome Dice')
dice_option_var = tk.StringVar(value='White Dice')
dice_size_var = tk.StringVar(value='Standard Dice (0.625")')
image_path_var = tk.StringVar()

# Color selection variables
color_vars = {
    'white': tk.BooleanVar(value=True),
    'black': tk.BooleanVar(value=False),
    'red': tk.BooleanVar(value=False),
    'blue': tk.BooleanVar(value=False),
    'yellow': tk.BooleanVar(value=False)
}

# Global variables
last_output_img = None  # Stores the last generated image
preview_image_tk = None  # For the image on the canvas
dice_layout_text = ""  # Stores the dice layout text


def browse_image():
    """Function to browse and select an image file."""
    file_path = filedialog.askopenfilename(
        filetypes=[("Image files", "*.jpg;*.jpeg;*.png;*.bmp;*.gif")])
    if file_path:
        image_path_var.set(file_path)


def floyd_steinberg_dithering(img_array):
    """Applies Floyd-Steinberg dithering to the image array."""
    height, width = img_array.shape
    new_img = img_array.astype(float)

    for y in range(height):
        for x in range(width):
            old_pixel = new_img[y, x]
            new_pixel = np.round(old_pixel / 32) * 32  # Quantize to 8 levels (0-7)
            quant_error = old_pixel - new_pixel
            new_img[y, x] = new_pixel

            if x + 1 < width:
                new_img[y, x + 1] += quant_error * 7 / 16
            if x - 1 >= 0 and y + 1 < height:
                new_img[y + 1, x - 1] += quant_error * 3 / 16
            if y + 1 < height:
                new_img[y + 1, x] += quant_error * 5 / 16
            if x + 1 < width and y + 1 < height:
                new_img[y + 1, x + 1] += quant_error * 1 / 16

    # Normalize new_img to 0-7
    dice_values = np.floor(new_img / 32).astype(int)
    dice_values = np.clip(dice_values, 0, 7)
    return dice_values


def map_grayscale_to_colors(dice_values, selected_colors):
    """Map dice values to colors based on grayscale intensity."""
    # Define perceived brightness for each color
    color_brightness = {
        'black': 0,
        'red': 76,
        'blue': 29,
        'yellow': 225,
        'white': 255
    }

    # Filter brightness for selected colors
    selected_brightness = {color: color_brightness[color] for color in selected_colors}
    # Sort colors by brightness
    sorted_colors = sorted(selected_brightness, key=selected_brightness.get)
    num_colors = len(sorted_colors)

    # Map dice values to color indices
    color_indices = (dice_values / 7 * (num_colors - 1)).astype(int)
    color_indices = np.clip(color_indices, 0, num_colors - 1)

    # Create a color map
    color_map = {i: sorted_colors[i] for i in range(num_colors)}
    # Map to colors
    color_mapped_values = np.vectorize(color_map.get)(color_indices)
    return color_mapped_values


def generate_dice_image():
    """Function to generate the dice image based on user inputs."""
    global last_output_img, dice_layout_text
    try:
        # Get user inputs
        physical_width_ft = physical_width_var.get()
        physical_height_ft = physical_height_var.get()
        dice_type = dice_type_var.get()
        dice_option = dice_option_var.get()
        dice_size_option = dice_size_var.get()
        image_path = image_path_var.get()
        selected_colors = [color for color, var in color_vars.items() if var.get()]

        # Validate inputs
        if not image_path:
            messagebox.showerror("Error", "Please select an image file.")
            return

        if physical_width_ft <= 0 or physical_height_ft <= 0:
            messagebox.showerror("Error", "Please enter valid physical dimensions.")
            return

        if dice_type == 'Colored Dice' and not selected_colors:
            messagebox.showerror("Error", "Please select at least one dice color.")
            return

        # Set dice size based on the selected option
        if dice_size_option == 'Standard Dice (0.625")':
            DICE_SIZE_INCHES = 0.625
        elif dice_size_option == 'Mini Dice (0.27")':
            DICE_SIZE_INCHES = 0.27
        elif dice_size_option == 'Micro Dice (0.19685")':
            DICE_SIZE_INCHES = 0.19685  # 5 mm in inches
        else:
            messagebox.showerror("Error", "Invalid dice size selected.")
            return

        # Convert physical dimensions to inches
        physical_width_in = physical_width_ft * INCHES_PER_FOOT
        physical_height_in = physical_height_ft * INCHES_PER_FOOT

        # Calculate the number of dice horizontally and vertically
        num_dice_horizontal = int(round(physical_width_in / DICE_SIZE_INCHES))
        num_dice_vertical = int(round(physical_height_in / DICE_SIZE_INCHES))

        # Load the image
        img = Image.open(image_path)
        image_width, image_height = img.size

        # Calculate aspect ratios
        image_aspect_ratio = image_width / image_height
        desired_aspect_ratio = physical_width_in / physical_height_in

        # Compare aspect ratios
        tolerance = 0.1  # Allowable difference in aspect ratio
        if abs(image_aspect_ratio - desired_aspect_ratio) > tolerance:
            proceed = messagebox.askyesno(
                "Aspect Ratio Mismatch",
                "The aspect ratio of the image does not match the desired physical dimensions. Do you want to proceed anyway?"
            )
            if not proceed:
                return

        # Convert image to grayscale
        img = img.convert('L')

        # Resize the image
        img = img.resize((num_dice_horizontal, num_dice_vertical), Image.LANCZOS)

        # Convert image to numpy array
        img_array = np.array(img)

        # Apply dithering
        dice_values = floyd_steinberg_dithering(img_array)

        if dice_type == 'Colored Dice':
            # Map grayscale values to selected colors
            dice_colors = map_grayscale_to_colors(dice_values, selected_colors)
        elif dice_option == 'Combined Dice':
            # Map grayscale values to black and white based on thresholds
            dice_colors = np.where(dice_values <= 3, 'black', 'white')
        else:
            dice_colors = np.full(dice_values.shape, dice_option.lower().split()[0])  # e.g., 'white' or 'black'

        # Generate the dice image preview
        output_img, dice_faces = create_dice_image(
            dice_values, dice_type, dice_option, selected_colors,
            num_dice_horizontal, num_dice_vertical, dice_colors
        )
        if output_img is None:
            return  # Error occurred

        last_output_img = output_img  # Store the generated image

        # Update the preview in the GUI
        display_preview(output_img)

        # Build the dice layout and total number of dice
        dice_layout_text = build_layout_and_count(
            dice_colors, dice_faces,
            num_dice_horizontal, num_dice_vertical,
            dice_type, selected_colors
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))


def create_dice_image(
        dice_values, dice_type, dice_option, selected_colors,
        num_dice_horizontal, num_dice_vertical, dice_colors
):
    """Function to create the dice image based on the dice values and options."""
    dice_face_size_px = 16  # Adjust size for GUI display
    dice_imgs = {}
    dice_faces = np.empty(dice_values.shape, dtype=object)

    try:
        if dice_type == 'Monochrome Dice':
            if dice_option == 'White Dice':
                dice_color = 'white'
            elif dice_option == 'Black Dice':
                dice_color = 'black'
            elif dice_option == 'Combined Dice':
                # Load both white and black dice images
                dice_imgs['black'] = {}
                dice_imgs['white'] = {}
                # Load dice face images
                for i in range(1, 7):
                    dice_imgs['black'][str(i)] = Image.open(f'dice_black/{i}.png').resize(
                        (dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                    dice_imgs['white'][str(i)] = Image.open(f'dice_white/{i}.png').resize(
                        (dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                # Load solid dice images
                dice_imgs['black']['solid'] = Image.open('dice_black/solid_black.png').resize(
                    (dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                dice_imgs['white']['solid'] = Image.open('dice_white/solid_white.png').resize(
                    (dice_face_size_px, dice_face_size_px), Image.LANCZOS)
            else:
                messagebox.showerror("Error", "Invalid dice option selected.")
                return None, None

            if dice_option != 'Combined Dice':
                # Load dice face images for single color
                dice_imgs = {}
                for i in range(1, 7):
                    path = f"dice_{dice_color}/{i}.png"
                    dice_img = Image.open(path).resize((dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                    dice_imgs[str(i)] = dice_img

                # Load solid dice images
                dice_imgs['solid'] = Image.open(f'dice_{dice_color}/solid_{dice_color}.png').resize(
                    (dice_face_size_px, dice_face_size_px), Image.LANCZOS)

        elif dice_type == 'Colored Dice':
            # Load dice images for selected colors
            for color in selected_colors:
                # Load dice face images
                color_imgs = {}
                for i in range(1, 7):
                    path = f'dice_{color}/{i}.png'
                    dice_img = Image.open(path).resize((dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                    color_imgs[str(i)] = dice_img
                # Load solid dice image
                solid_path = f'dice_{color}/solid_{color}.png'
                solid_img = Image.open(solid_path).resize((dice_face_size_px, dice_face_size_px), Image.LANCZOS)
                color_imgs['solid'] = solid_img
                # Store images for the color
                dice_imgs[color] = color_imgs
        else:
            messagebox.showerror("Error", "Invalid dice type selected.")
            return None, None

        # Create a new image to hold the dice layout
        output_img = Image.new('RGB', (num_dice_horizontal * dice_face_size_px, num_dice_vertical * dice_face_size_px))

        for y in range(num_dice_vertical):
            for x in range(num_dice_horizontal):
                if dice_type == 'Colored Dice':
                    color = dice_colors[y, x]
                    # Map grayscale intensity to dice face value (1-6)
                    dice_face_value = int(dice_values[y, x] % 6) + 1  # Values between 1 and 6
                    dice_face_img = dice_imgs[color][str(dice_face_value)]
                    dice_faces[y, x] = f"{color[:3].capitalize()}{dice_face_value}"
                else:
                    dice_value = dice_values[y, x]
                    color = dice_colors[y, x]
                    if dice_option == 'Combined Dice':
                        # Map values to images
                        if dice_value == 0 or dice_value == 7:
                            dice_face_img = dice_imgs[color]['solid']
                            dice_faces[y, x] = f"{color[:3].capitalize()}S"
                        else:
                            dice_face_img = dice_imgs[color][str(dice_value)]
                            dice_faces[y, x] = f"{color[:3].capitalize()}{dice_value}"
                    else:
                        if dice_value == 0 or dice_value == 7:
                            dice_face_img = dice_imgs['solid']
                            dice_faces[y, x] = f"{dice_color[:3].capitalize()}S"
                        else:
                            dice_face_img = dice_imgs[str(dice_value)]
                            dice_faces[y, x] = f"{dice_color[:3].capitalize()}{dice_value}"

                output_img.paste(dice_face_img, (x * dice_face_size_px, y * dice_face_size_px))

        return output_img, dice_faces

    except Exception as e:
        messagebox.showerror("Error", f"Error creating dice image: {str(e)}")
        return None, None


def display_preview(image):
    """Function to display the preview image in the GUI."""
    global preview_image_tk  # Keep a reference to prevent garbage collection

    # Resize the image to fit within the canvas while maintaining aspect ratio
    canvas_width = preview_canvas.winfo_width()
    canvas_height = preview_canvas.winfo_height()

    if canvas_width == 1 and canvas_height == 1:
        # Canvas size not initialized yet, set default size
        canvas_width = 500
        canvas_height = 500

    # Calculate the scaling factor
    img_width, img_height = image.size
    scale = min(canvas_width / img_width, canvas_height / img_height)

    new_size = (int(img_width * scale), int(img_height * scale))
    preview_img = image.resize(new_size, Image.LANCZOS)

    # Create ImageTk object
    preview_image_tk = ImageTk.PhotoImage(preview_img)

    # Clear the canvas
    preview_canvas.delete('all')

    # Center the image on the canvas
    x = (canvas_width - new_size[0]) // 2
    y = (canvas_height - new_size[1]) // 2

    # Add the image to the canvas
    preview_canvas.create_image(x, y, anchor='nw', image=preview_image_tk)


def build_layout_and_count(
        dice_colors, dice_faces, num_dice_horizontal, num_dice_vertical,
        dice_type, selected_colors
):
    """Function to build the dice layout text and count dice by color."""
    # Display total number of dice needed
    total_dice = num_dice_horizontal * num_dice_vertical

    # Count the number of dice for each color
    color_counts = {}
    if dice_type == 'Colored Dice':
        unique_colors, counts = np.unique(dice_colors, return_counts=True)
        for color, count in zip(unique_colors, counts):
            color_counts[color] = count
    else:
        # For monochrome dice
        if dice_option_var.get() == 'Combined Dice':
            unique_colors, counts = np.unique(dice_colors, return_counts=True)
            for color, count in zip(unique_colors, counts):
                color_counts[color] = count
        else:
            color = dice_option_var.get().split()[0].lower()
            color_counts[color] = total_dice

    # Build the total dice text
    total_dice_text = f"Total number of dice needed: {total_dice}\n"
    for color, count in color_counts.items():
        color_name = color.capitalize()
        total_dice_text += f"  {color_name} dice: {count}\n"

    # Build the dice layout grid
    layout_lines = []

    # Header row with column numbers
    header = "    |" + "|".join(f"{col:>5}" for col in range(1, num_dice_horizontal + 1)) + "|"
    separator = "----" + "+-----" * num_dice_horizontal + "+"
    layout_lines.append(header)
    layout_lines.append(separator)

    for idx, row in enumerate(dice_faces):
        row_num = f"{idx + 1:>3} |"
        row_values = "|".join(f"{val:>5}" for val in row)
        layout_lines.append(f"{row_num}{row_values}|")
        layout_lines.append(separator)

    layout_text = "\n".join(layout_lines)

    # Add a legend at the end
    legend = "\nLegend:\n"
    if dice_type == 'Colored Dice':
        for color in selected_colors:
            legend += f"{color[:3].capitalize()} - {color.capitalize()} Dice\n"
    else:
        if dice_option_var.get() == 'Combined Dice':
            legend += "Bla - Black Dice\n"
            legend += "Whi - White Dice\n"
            legend += "S - Solid Face\n"
            legend += "1 to 6 - Dice Face with that Number\n"
        else:
            color = dice_option_var.get().split()[0]
            legend += f"{color[:3]} - {color} Dice\n"
            legend += "S - Solid Face\n"
            legend += "1 to 6 - Dice Face with that Number\n"

    layout_text += legend

    # Update the total dice label
    total_dice_label.config(text=total_dice_text)

    return layout_text


def save_layout():
    """Function to save the dice layout to a text file."""
    if not dice_layout_text:
        messagebox.showerror("Error", "No layout to save. Please generate an image first.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension='.txt', filetypes=[('Text files', '*.txt')])
    if file_path:
        with open(file_path, 'w') as file:
            file.write(dice_layout_text)
        messagebox.showinfo("Success", f"Layout saved to {file_path}")


def save_layout_to_excel():
    """Function to save the dice layout to an Excel file."""
    if not dice_layout_text:
        messagebox.showerror("Error", "No layout to save. Please generate an image first.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        try:
            wb = Workbook()
            ws = wb.active

            # Split the dice_layout_text into lines
            layout_text = dice_layout_text.splitlines()

            # Skip the legend at the end
            # Find the line where the legend starts
            for idx, line in enumerate(layout_text):
                if line.startswith("Legend:"):
                    legend_start = idx
                    break
            else:
                legend_start = len(layout_text)

            # Extract the grid lines
            grid_lines = layout_text[0:legend_start]

            # Parse the grid and write to Excel
            row_counter = 1
            for line in grid_lines:
                # Skip separator lines
                if line.startswith("----") or line.strip() == '':
                    continue
                # Parse header row
                if line.startswith("    |"):
                    # Extract column numbers
                    columns = line.split('|')[1:-1]
                    for col_idx, col_num in enumerate(columns, start=2):  # Start from column B
                        ws.cell(row=row_counter, column=col_idx, value=col_num.strip())
                    row_counter += 1
                else:
                    # Extract row data
                    parts = line.split('|')
                    row_num = parts[0].strip()
                    if row_num == '':
                        continue
                    ws.cell(row=row_counter, column=1, value=row_num)  # Row number in column A
                    values = parts[1:-1]
                    for col_idx, val in enumerate(values, start=2):
                        ws.cell(row=row_counter, column=col_idx, value=val.strip())
                    row_counter += 1

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Layout saved to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the layout: {e}")


def save_preview_image():
    """Function to save the preview image to a file."""
    if last_output_img is None:
        messagebox.showerror("Error", "No image to save. Please generate an image first.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension='.png', filetypes=[
        ('PNG files', '*.png'),
        ('JPEG files', '*.jpg;*.jpeg'),
        ('Bitmap files', '*.bmp'),
        ('GIF files', '*.gif')
    ])
    if file_path:
        try:
            # Save the last generated image
            last_output_img.save(file_path)
            messagebox.showinfo("Success", f"Preview image saved to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the image: {e}")


def update_color_options(*args):
    # Show or hide color options based on dice type
    if dice_type_var.get() == 'Colored Dice':
        color_frame.grid(row=4, column=0, columnspan=3, padx=5, pady=5)
        dice_option_label.grid_remove()
        dice_option_menu.grid_remove()  # Hide monochrome dice option
    else:
        color_frame.grid_remove()
        dice_option_label.grid()
        dice_option_menu.grid()  # Show monochrome dice option


# Frame for inputs
input_frame = ttk.Frame(main_frame)
input_frame.pack(pady=10, padx=10)

# Physical Width
width_label = ttk.Label(input_frame, text="Width (ft):")
width_label.grid(row=0, column=0, padx=5, pady=5, sticky='e')
width_entry = ttk.Entry(input_frame, textvariable=physical_width_var)
width_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

# Physical Height
height_label = ttk.Label(input_frame, text="Height (ft):")
height_label.grid(row=1, column=0, padx=5, pady=5, sticky='e')
height_entry = ttk.Entry(input_frame, textvariable=physical_height_var)
height_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')

# Image Selection
image_label = ttk.Label(input_frame, text="Image File:")
image_label.grid(row=2, column=0, padx=5, pady=5, sticky='e')
image_entry = ttk.Entry(input_frame, textvariable=image_path_var, width=40)
image_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
browse_button = ttk.Button(input_frame, text="Browse", command=browse_image)
browse_button.grid(row=2, column=2, padx=5, pady=5, sticky='w')

# Dice Type Option
dice_type_label = ttk.Label(input_frame, text="Dice Type:")
dice_type_label.grid(row=3, column=0, padx=5, pady=5, sticky='e')
dice_type_menu = ttk.OptionMenu(
    input_frame,
    dice_type_var,
    dice_type_var.get(),
    'Monochrome Dice',
    'Colored Dice'
)
dice_type_menu.grid(row=3, column=1, padx=5, pady=5, sticky='w')

# Dice Option Menu (for Monochrome Dice)
dice_option_label = ttk.Label(input_frame, text="Dice Option:")
dice_option_label.grid(row=4, column=0, padx=5, pady=5, sticky='e')
dice_option_menu = ttk.OptionMenu(
    input_frame,
    dice_option_var,
    dice_option_var.get(),
    'White Dice',
    'Black Dice',
    'Combined Dice'
)
dice_option_menu.grid(row=4, column=1, padx=5, pady=5, sticky='w')

# Color Selection Frame (Initially Hidden)
color_frame = ttk.Frame(input_frame)

# Create checkboxes for each color
color_label = ttk.Label(color_frame, text="Select Colors:")
color_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='n')

for idx, (color, var) in enumerate(color_vars.items()):
    col = idx % 2
    row = idx // 2 + 1
    cb = ttk.Checkbutton(color_frame, text=color.capitalize(), variable=var)
    cb.grid(row=row, column=col, sticky='w', padx=5, pady=2)

# Now set up the trace and call update_color_options
dice_type_var.trace('w', update_color_options)
update_color_options()

# Dice Size Menu
dice_size_label = ttk.Label(input_frame, text="Dice Size:")
dice_size_label.grid(row=5, column=0, padx=5, pady=5, sticky='e')
dice_size_menu = ttk.OptionMenu(
    input_frame,
    dice_size_var,
    dice_size_var.get(),
    'Standard Dice (0.625")',
    'Mini Dice (0.27")',
    'Micro Dice (0.19685")'  # 5 mm dice
)
dice_size_menu.grid(row=5, column=1, padx=5, pady=5, sticky='w')

# Generate Button
generate_button = ttk.Button(main_frame, text="Generate Dice Image", command=generate_dice_image)
generate_button.pack(pady=10)

# Create a frame for the preview image
preview_frame = ttk.Frame(main_frame)
preview_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Create the canvas for the preview image
preview_canvas = tk.Canvas(preview_frame, bg='white')
preview_canvas.pack(fill=tk.BOTH, expand=True)

# Total Dice Label
total_dice_label = ttk.Label(main_frame, text="")
total_dice_label.pack(pady=5)

# Information Label
info_label = ttk.Label(main_frame, text="You can save the dice layout grid as a text or Excel file.")
info_label.pack(pady=5)

# Save Layout Buttons Frame
buttons_frame = ttk.Frame(main_frame)
buttons_frame.pack(pady=10)

save_layout_button = ttk.Button(buttons_frame, text="Save Layout as Text", command=save_layout)
save_layout_button.grid(row=0, column=0, padx=5, pady=5)

save_layout_excel_button = ttk.Button(buttons_frame, text="Save Layout to Excel", command=save_layout_to_excel)
save_layout_excel_button.grid(row=0, column=1, padx=5, pady=5)

save_image_button = ttk.Button(buttons_frame, text="Save Preview Image", command=save_preview_image)
save_image_button.grid(row=0, column=2, padx=5, pady=5)

# Center buttons
buttons_frame.columnconfigure(0, weight=1)
buttons_frame.columnconfigure(1, weight=1)
buttons_frame.columnconfigure(2, weight=1)

# Start the GUI main loop
root.mainloop()
